# python3!

import openpyxl


'''
每一行代表一次单独的销售。列分别是销售产品的类型（A） 、产品每磅的价格
（B） 、销售的磅数（C） ，以及这次销售的总收入。TOTAL 列设置为 Excel 公式，将
每磅的成本乘以销售的磅数，并将结果取整到分。有了这个公式，如果列 B 或 C 发
生变化，TOTAL 列中的单元格将自动更新。 
现在假设 Garlic、 Celery 和 Lemons 的价格输入的不正确。 这让你面对一项无聊
的任务：遍历这个电子表格中的几千行，更新所有 garlic、celery 和 lemon 行中每磅
的价格。你不能简单地对价格查找替换，因为可能有其他的产品价格一样，你不希
望错误地“更正” 。对于几千行数据，手工操作可能要几小时。但你可以编写程序，
几秒钟内完成这个任务。 
你的程序做下面的事情： 
•  循环遍历所有行。 
•  如果该行是 Garlic、Celery 或 Lemons，更新价格。 
这意味着代码需要做下面的事情： 
•  打开电子表格文件。 
•  针对每一行，检查列 A 的值是不是 Celery、Garlic 或 Lemon。 
•  如果是，更新列 B 中的价格。 
•  将该电子表格保存为一个新文件（这样就不会丢失原来的电子表格，以防万一） 。
文件名：produceSales.xlsx
文件格式如下： 
PRODUCE	    COST PER POUND	    POUNDS SOLD	    TOTAL
Potatoes	    0.86	         21.6	        18.58
'''
'''
需要更新的价格如下： 
Celery  1.19 
Garlic  3.07 
Lemon  1.27 
'''

newCost = {
    "Celery" : 1.19,
    "Garlic" : 3.07,
    "Lemon" : 1.27
    }

# •  打开电子表格文件。
wb = openpyxl.load_workbook('produceSales.xlsx')
sh = wb['Sheet']
# •  针对每一行，检查列 A 的值是不是 Celery、Garlic 或 Lemon。
# for row in range(2,sh.max_row+1):
#     ProductName = sh['a' + str(row)].value
#     newCost.setdefault(ProductName,0)
#     # •  如果是，更新列 B 中的价格。
#     if newCost[ProductName] :
#         sh['b' + str(row)] = newCost[ProductName]


# 优化后代码
for row in range(2,sh.max_row+1):
    ProductName = sh['a'+str(row)].value
    if ProductName in newCost:
        sh['b' + str(row)] = newCost[ProductName]
        print(sh['a' + str(row)].value,sh['b' + str(row)].value,sh['c' + str(row)].value,sh['d' + str(row)].value)

# •  将该电子表格保存为一个新文件（这样就不会丢失原来的电子表格，以防万一） 。

wb.save('newProduceSales.xlsx')
wb.close()